import openpyxl
import random
import sys
import regex as re
from defs import scrape
from defs import dieRoll


## TODO: webscraper til https://dungeonmastertools.github.io/index.html
# web scraping unit


# Introduserer Excel
wb = openpyxl.load_workbook('C:/Users/Bruker/Documents/Rollespill/D&D/DMGenerator/Treasure/TreasureChart.xlsx')
sheet = wb.sheetnames[0]


def treasure(type, cr):
    treasureRoll = random.randrange(1, 101)  # diceroll

    diceString = ""  # Kun for å fikse feilmelding
    # Henter ut streng fra arket
    if type == "loot":
        diceString = lootRoll(cr, treasureRoll)
    elif type == "hoard":
        diceString = hoardRoll(cr, treasureRoll)
    else:
        print("Galt input: treasure() -- Ikke \"loot\" eller \"hoard\".")
        exit(2)

    #print(diceString)


def findCell(cr, treasureRoll, rangemax):
    # Introduserer Excel
    # wb = openpyxl.load_workbook('Treasure/TreasureChart.xlsx')
    # sheet = wb.sheetnames[0]
    column = cr

    for i in range(1, rangemax + 1):
        tryTable = wb[sheet].cell(i, column).value
        inBetweens = tryTable.split(",")
        lower = int(inBetweens[0])
        upper = int(inBetweens[1])
        if treasureRoll in range(lower, upper):
            lootDiceString = wb[sheet].cell(i, column + 1).value
            return lootDiceString
    print("findCell fant ikke riktig rute")
    exit(2)


def findCellHoard(cr, treasureRoll):
    column = cr

    lootDiceString = wb[sheet].cell(treasureRoll, column + 1).value
    return lootDiceString


def rollDice(rollAmount, rollSize, multiplier):
    totalt = 0
    for i in range(rollAmount):
        totalt += random.randint(1, rollSize)
    totalt = totalt * multiplier
    return totalt


def equationSeparatorCash(characters):
    # string = KP:4d6x100;GP:1d6x5
    # Må fungere parallelt på ett og to elementer. Noen har to deler med semikolon-separator, andre har ikke.

    valører = []
    antTerninger = []
    terningtyper = []
    multi = []
    totalListe = []

    mynt_sep = characters.split(";")
    #print(mynt_sep)
    for i in mynt_sep:
        valør = re.search(r"[A-Z]{2}",i).span()
        valører.append(i[valør[0]:valør[1]])

        antTerning = re.search(r"\d",i).span()
        antTerninger.append(i[antTerning[0]:antTerning[1]])

        terningtype = list(re.search(r"d\d",i).span())
        terningtype[0] += 1
        terningtype = tuple(terningtype)
        terningtyper.append(i[terningtype[0]:terningtype[1]])

        multien = list(re.search(r"x\d{1,4}",i).span())
        multien[0] += 1
        multien = tuple(multien)
        multi.append(i[multien[0]:multien[1]])

    while len(valører) > 0:
        pengeMengde = rollDice(int(antTerninger.pop(0)), int(terningtyper.pop(0)), int(multi.pop(0)))
        totalListe.append(str(pengeMengde) + " " + valører.pop(0))

    svarStreng = "De finner "
    for e in range(len(totalListe)):
        svarStreng = svarStreng + totalListe[e]
        if len(totalListe) - e == 1:
            pass
        elif len(totalListe) - e == 2:
            svarStreng = svarStreng + " og "
        else:
            svarStreng = svarStreng + ", "
    return svarStreng


def equationSeparatorMagic(string):
    # string =  1d6;C, 1d4;A,1d6;B eller ""
    if string is None:
        return "Ingen magiske gjenstander"

    if "," in string:
        string = string.split(",")
    else:
        string = [string]

    typer = len(string)

    typeItems = []
    antTerninger = []
    typeTerning = []
    totalListe = []

    for type in range(typer):
        typeItems.append(string[type].split(";")[1])  # A
        rest = string[type].split(";")[0]  # 1d6

        antTerninger.append(int(rest.split("d")[0]))  # 1
        typeTerning.append(int(rest.split("d")[1]))  # 6

        antGjenstander = rollDice(antTerninger[type], typeTerning[type], 1)
        if antGjenstander == 1:
            totalListe.append(
            str(antGjenstander) + str(typeItems[0]))
        else:
            totalListe.append(
            str(antGjenstander) + str(typeItems[0]))

    findMagic(totalListe)
    ##Alt under her er borked!!!!!!

    svarStreng = ""
    for e in range(len(totalListe)):
        svarStreng = svarStreng + totalListe[e]
        if len(totalListe) - e == 1:
            pass
        elif len(totalListe) - e == 2:
            svarStreng = svarStreng + " og "
        else:
            svarStreng = svarStreng + ", "



    return svarStreng


def findMagic(string):
    #print(string)
    # string = ["5A", "1C", "3G"]
    if len(string) < 1:
        return

    tables = scrape("https://dungeonmastertools.github.io/index.html", "tbody", {"class":"table-hover"})
    #for t in len(tables):
    #    tables[t] = tables[t]{"tbody class = \"table"}
    #print(tables[0])


    tables_indices_lookup = {"A" : 1,
                    "B" : 2,
                    "C" : 3,
                    "D" : 4,
                    "E" : 5,
                    "F" : 6,
                    "G" : 7,
                    "H" : 8,
                    "I" : 9}

    amount_index_tuple = []
    for element in string:
        amount = int(element[0])
        item = element[1]
        item_index = tables_indices_lookup[item]
        amount_index_tuple.append((amount, item_index))
    
    def evacuate(string, substring_start, substring_end):
        print(type(string))
        print(string.find(substring_start))
        print(len(substring_start))
        evacuate_from = string.find(substring_start) + len(substring_start)
        evacuate_to = string.find(substring_end) - 1

        evacuated_value = string[evacuate_from, evacuate_to]
        evacuated_value.replace("–", ",")
        evacuated_value.replace("00", "100")

        string.replace(substring_start, "", 1)# Remove used piece of string
        string.replace(substring_end, "", 1)# Remove one instance of substring, not all

        if evacuate_to == -1 or evacuate_from == -1:
            return -1
        return evacuated_value
    
    # evac = evacuate(tables)

    number_evacuated = evacuate(table, "<td>", "</td>")
    item_evacuated = evacuate(table, "<td>", "</td>")
    numbers_and_items_string_pair = []
    for table in tables:
        string_list = table.find_all("tbody", "table-hover")

        
        while number_evacuated != -1 and item_evacuated != -1:
            numbers_and_items_string_pair.append((number_evacuated, item_evacuated))
            number_evacuated = evacuate(table, "<td>", "</td>")
            item_evacuated = evacuate(table, "<td>", "</td>")
        
    
    print(f"{string_list=}")
    for number, item in string_list:
        print(f"{number=},{item}")
        number = number.split()[0]# Remove whitespace
        if len(number) > 2:
            number = (int(number[:1]), int(number[-2:]))
        elif number == "00":
            number = (100,)
        else:
            number = (int(number),)

    for item_type in amount_index_tuple:
        for amount in range(item_type[0]):
            die_number = dieRoll(100)



def lootRoll(cr, treasureRoll):
    lootCRDict = {
        range(0, 5): 1,
        range(5, 11): 3,
        range(11, 17): 5,
        range(17, sys.maxsize): 7
    }
    columnForCR = rangeIter(lootCRDict, cr)

    lootDice = findCell(columnForCR, treasureRoll, 5)
    svarListe = equationSeparatorCash(lootDice)
    return svarListe


def hoardRoll(cr, treasureRoll):
    hoardDiceDict = {
        range(0, 5): 9,
        range(5, 11): 12,
        range(11, 17): 15,
        range(17, sys.maxsize): 18
    }
    columnForCR = rangeIter(hoardDiceDict, cr)

    hoardDiceCash = findCellHoard(columnForCR, treasureRoll)
    hoardCashString = equationSeparatorCash(hoardDiceCash)  # string

    hoardDiceObjects = findCellHoard(columnForCR, treasureRoll)  # 2d6x10gp
    hoardObjectString = equationSeparatorCash(hoardDiceObjects)  # string

    objectType = findCellHoard(columnForCR + 1, treasureRoll)  # string

    hoardDiceMagic = findCellHoard(columnForCR + 2, treasureRoll)
    hoardMagicString = equationSeparatorMagic(hoardDiceMagic)  # string

    svarStreng = f"{hoardCashString}\n{hoardObjectString} {objectType}\n{hoardMagicString}"

    return svarStreng


def rangeIter(iterable, value):
    for element in iterable:
        if value in element:
            return iterable[element]
    print(f"Value out of range: rangeIter. Value: {value}")
    # exit(2)
