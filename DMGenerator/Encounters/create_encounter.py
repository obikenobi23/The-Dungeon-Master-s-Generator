import openpyxl
import math
from Encounters.Encounter import encounter
from Encounters.findMonster import findMonster



def createEncounter():
    # Initialisér verdier
    diffBok = openpyxl.load_workbook("Encounters/DiffBok.xlsx")
    dSheet = diffBok.active
    crSpm = input("Hvilken vanskegrad vil du ha? ")

    def crSpm_filter(cr):
        if cr.lower() in {"e", "easy"}:
            cr = "Easy"
        elif cr.lower() in {"m", "medium"}:
            cr = "Medium"
        elif cr.lower() in {"h", "hard"}:
            cr = "Hard"
        elif cr.lower() in {"d", "deadly"}:
            cr = "Deadly"
        return cr

    crSpm = crSpm_filter(crSpm)
    while crSpm not in {"Easy", "Medium", "Hard", "Deadly"}:
        crSpm = input("Skrivefeil; prøv igjen")
        crSpm = crSpm_filter(crSpm)

    def lvlListe(liste):
        while True:
            heroLvl = input("Legg inn levelet til en helt, trykk \"Enter\" for å gå videre\n")
            if heroLvl == "":
                return liste
            else:
                liste.append(int(heroLvl))

    # Let etter riktig vanskegrad i boken
    lvler = lvlListe([])
    teamMembers = len(lvler)
    avgLvl = math.floor(sum(lvler) / teamMembers)
    column = 1
    for colNum in range(2, 6):
        if crSpm == dSheet.cell(1, colNum).value:
            column = colNum  # Angi hvilken vanskegrad maksExp skal hentes fra
            break

    # Definér maksimal exp for encounteret
    maxExp = teamMembers * dSheet.cell(avgLvl + 1, column).value
    print(maxExp)

    # Biome defineres
    def findBiomeList(liste):
        biome = input("Hvilket miljø kjemper vi i? (enter for å avslutte) (Tom liste = kun \"dungeon\")")
        if biome != "":  # rekursér
            liste.append(biome)
            liste.append(findBiomeList(liste))
        else:  # avslutt
            if len(liste) == 0:
                liste.append("dungeon")
            #else:
                #liste.append(biome)
        return liste

    biomeList = findBiomeList([])

    # Kjøre findMonster og encounter med parameter
    findMonster_list = findMonster(biomeList)
    validMonsters = findMonster_list[0]
    validMonsters_cr = findMonster_list[1]

    # Over til deg, encounter()
    encounter(maxExp, validMonsters, validMonsters_cr)
