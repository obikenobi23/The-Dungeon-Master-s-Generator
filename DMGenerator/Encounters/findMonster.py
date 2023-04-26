## Generer en liste over akutelle monstre, samt tilhørende liste med cr
# Importer moduler
import openpyxl


def findMonster(biomeList):
    # Hent definisjoner fra hovedprogrammet
    monsterBok = openpyxl.load_workbook("Encounters/BookOfMonsters.xlsx")
    sheets = monsterBok.sheetnames

    # Lag en liste over alle vesnene som er i de riktige biomer, velg ut vesner fra den listen
    # Hent ut brukbare monstre og gjør dem tilgjengelige for videre regning
    validMonsters = []
    validMonsters_cr = []
    for sheet in range(len(sheets)):
        cr = monsterBok[sheets[sheet]]  # cr er det til enhver tid gjeldende ark

        for i in range(2, cr.max_row + 1):  # Tell gjennom rekkene der monstre er lagret i arket

            # Lag en liste av hvert monsters naturlige miljø
            biomePreference = str(cr.cell(i, 3).value)
            biomePreference = biomePreference.split(",")

            for j in biomePreference:  # Sammenlikn data fra den gjeldene liste med brukergitte parameter om biomer
                if j in biomeList:
                    if cr.cell(i, 2).value not in validMonsters:
                        validMonsters.append(str(cr.cell(i, 2).value))
                        validMonsters_cr.append(str(sheet))

    return validMonsters, validMonsters_cr
