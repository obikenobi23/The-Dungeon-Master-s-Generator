# Her genereres feller og annet artig
import openpyxl
import random


def trap(type):
    # Opprett kontakt med excel
    wb = openpyxl.load_workbook("DungeonGenerator/DungeonParts.xlsx")
    sheet = wb.sheetnames[5]

    # For traps
    if type == "trap":

        # Angi randomisering
        trigger_diceRoll = random.randrange(1, 7)
        sever_diceRoll = random.randrange(1, 7)
        effect_diceRoll = random.randrange(1, 101)

        # Hent randomisert verdi fra riktig kolonne
        trigger = wb[sheet].cell(trigger_diceRoll, 1).value.lower()
        severity = wb[sheet].cell(sever_diceRoll, 2).value.lower()
        effect = wb[sheet].cell(effect_diceRoll, 3).value.lower()

        # Skriv ut resultat
        print("The trap is {}. If you {} it, {}.".format(severity, trigger, effect))

    # For tricks
    elif type == "trick":

        # Ang randomisering
        object_diceRoll = random.randrange(1, 21)
        effect_diceRoll = random.randrange(1, 101)

        # Hent randomisert verdi fra riktig kolonne
        item = (wb[sheet].cell(object_diceRoll, 4).value).lower()
        effect = (wb[sheet].cell(effect_diceRoll, 5).value).lower()

        # Skriv ut resultat
        print("The trick is a {} which {}.".format(item, effect))

    elif type == "hazard":

        # Angi randomisering
        diceRoll = random.randrange(1, 21)

        # Hent randomisert verdi fra riktig kolonne
        hazard = (wb[sheet].cell(diceRoll, 6).value).lower()

        # Skriv ut resultat
        print("The hazard is {}.".format(hazard))

    elif type == "obstacle":

        # Angi randomisering
        diceRoll = random.randrange(1, 21)

        # Hent randomisert verdi fra riktig kolonne
        obstacle = (wb[sheet].cell(diceRoll, 7).value).lower()

        # Skriv ut resultat
        print("The obstacle is {}.".format(obstacle))

    else:  # Feilmelding
        print("U fkd øp")
