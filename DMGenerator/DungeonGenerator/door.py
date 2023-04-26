import openpyxl
import random

def door():
    # Opprett kontakt med excel
    wb = openpyxl.load_workbook("DungeonGenerator/DungeonParts.xlsx")
    sheet = wb.sheetnames[1]

    # Hent en tilfeldig dør fra tabellen
    diceRoll = random.randrange(1, 21)
    pickedDoor = (wb[sheet].cell(diceRoll, 1).value).lower()
    print("The door is {}.".format(pickedDoor))

    # Lag en streng av dørens beskrivelse
    diceRoll = random.randrange(1, 21)
    beyond = wb[sheet].cell(diceRoll, 2).value
    print("Behind the door is a {}.".format(beyond))
