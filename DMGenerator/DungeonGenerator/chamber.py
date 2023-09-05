## Her genereres rom
import openpyxl
import random
from DungeonGenerator.door import door
from Encounters.create_encounter import createEncounter
from defs import waitRead


def roomFlavour(noise=False, air=False, odor=False):
    # Introduserer Excel
    wb = openpyxl.load_workbook("DungeonGenerator/DungeonParts.xlsx")
    sheet = wb.sheetnames[6]

    if noise:
        noise_diceRoll = random.randint(1, 101)
        noiseFlavour = wb[sheet].cell(noise_diceRoll, 1).value
        print("As you enter the room, you can hear {}.".format(noiseFlavour))

    if air:
        air_diceRoll = random.randint(1, 101)
        airFlavour = wb[sheet].cell(air_diceRoll, 2).value
        print("The air in the room is {}.".format(airFlavour))

    if odor:
        odor_diceRoll = random.randint(1, 101)
        odorFlavour = wb[sheet].cell(odor_diceRoll, 3).value
        print("The room smells {}.".format(odorFlavour))


def chamber():
    # Opprett kontakt med excel
    wb = openpyxl.load_workbook('DungeonGenerator/DungeonParts.xlsx')
    sheet = wb.sheetnames[3]

    # Hent ut sternger fra arket, omformater til tekst
    shape_diceRoll = random.randrange(1, 21)
    roomShape = wb[sheet].cell(shape_diceRoll, 1).value
    size_diceRoll = random.randrange(1, 21)
    roomSize = wb[sheet].cell(size_diceRoll, 2).value
    doorNum_diceRoll = random.randrange(1, 21)
    # Forskjellige tabeller for forskjellig størrelse på ormmet
    doorNum = 0
    if roomSize == "Normal":
        doorNum = wb[sheet].cell(doorNum_diceRoll, 3).value
    elif roomSize == "Large":
        doorNum = wb[sheet].cell(doorNum_diceRoll, 4).value
    contents_diceRoll = random.randrange(1, 101)
    contents = wb[sheet].cell(contents_diceRoll, 7).value

    # Vis bruker det genererte rom
    print(f"The room is a {roomShape}.")
    print(f"It contains {contents}.")

    # Generer dører automatisk
    for i in range(doorNum):
        exitType_diceRoll = random.randrange(1, 21)
        exitType = wb[sheet].cell(exitType_diceRoll, 6).value
        exitLocation_diceRoll = random.randrange(1, 21)
        exitLocation = wb[sheet].cell(exitLocation_diceRoll, 5).value.lower()
        print("There is a {} on the {}".format(exitType, exitLocation))
        if "door" in exitType.lower():
            door()

    waitRead(roomShape+contents)

    # Start generering av rommiljø
    flavourDice = random.randrange(1, 5)
    noise = False
    air = False
    odor = False
    if flavourDice > 1:
        noiseDice = random.randrange(1, 5)
        airDice = random.randrange(1, 5)
        odorDice = random.randrange(1, 5)

        if noiseDice == 4:
            noise = True
        if airDice == 4:
            air = True
        if odorDice == 4:
            odor = True

        roomFlavour(noise, air, odor)

    if "Monsters" in contents:
        createEncounter()
