import openpyxl
import random
from defs import waitRead

def passage():
    # Opprett kontakt med excel
    wb = openpyxl.load_workbook("DungeonGenerator/DungeonParts.xlsx")
    sheet = wb.sheetnames[2]

    # Hent ut strenger fra arket
    type_diceRoll = random.randrange(1,21)
    passageText = wb[sheet].cell(type_diceRoll,1).value
    width_diceRoll = random.randrange(1,21)
    passageWidth = (wb[sheet].cell(width_diceRoll,2).value).strip(" ").split(",")
    passageType = ""

    # Angi typen passasje
    if "Continue" in passageText:
        passageType = "passage"
    elif "Chamber" in passageText:
        passageType = "chamber"
    elif "Stairs" in passageText:
        passageType = "stairs"

    # Lag strenger for de forskjellige elementene i tabellen
    if len(passageWidth) == 1:
        text = "{}. The {} is {} wide.".format(passageText, passageType, passageWidth[0])
        print(text)
    elif width_diceRoll in range(17,19):
        text = "{}. The {} is {}, {}.".format(passageText, passageType, passageWidth[0],passageWidth[1])
        print(text)
    elif width_diceRoll == 19:
        text = "{}. The {} is roughly {} and {}.".format(passageText, passageType, passageWidth[0], passageWidth[1])
        print(text)
    elif width_diceRoll == 20:
        text = "{}, roughly {} and {}. There is a {}".format(passageText, passageWidth[0], passageWidth[1], passageWidth[2])
        print(text)

    waitRead(text)
