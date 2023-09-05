# Her genereres startlokasjon
import openpyxl
import random
from defs import waitRead


def start():
    # Opprett kontakt med excel
    wb = openpyxl.load_workbook("DungeonGenerator/DungeonParts.xlsx")
    sheet = wb.sheetnames[0]

    # Hent ut en streng fra arket
    diceRoll = random.randrange(1, 11)
    pickedStart = wb[sheet].cell(diceRoll, 1)

    # Del opp strengen til presentable deler
    pickedStart_list = pickedStart.value.strip("").split(";")

    # Ang delene av stengen sine navn
    roomShape = pickedStart_list[0]
    roomSize = pickedStart_list[1]
    roomExits = pickedStart_list[2]
    fourLong = bool
    fiveLong = bool
    extraExit = ""
    roomFlavour = ""
    if len(pickedStart_list) == 4:
        fourLong = True
        extraExit = pickedStart_list[3]
    elif len(pickedStart_list) == 5:
        fiveLong = True
        roomFlavour = pickedStart_list[4]

    # Skriv ut resultatet
    if fourLong:
        text = "You find yourself in a {} {}. There is {}. There is a {} as well".format(roomSize, roomShape.lower(), roomExits, extraExit)
        print(text)
    elif fiveLong:
        text = "You find yourself in a {} {} {}. There is {}. There is a {} as well".format(roomSize, roomShape.lower(), roomFlavour, roomExits, extraExit)
        print(text)
    else:
        text = "You find yourself in a {} {}. There is {}.".format(roomSize, roomShape.lower(), roomExits)
        print(text)

    waitRead(text)
