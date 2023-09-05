import openpyxl
import random
from defs import waitRead


def stairs():
    # Opprett kontakt med excel
    wb = openpyxl.load_workbook("DungeonGenerator/DungeonParts.xlsx")
    sheet = wb.sheetnames[4]

    # Hent en streng fra tabellen og skriv den ut
    diceRoll = random.randrange(1, 21)
    stairType = wb[sheet].cell(diceRoll, 1).value.lower()
    if diceRoll in range(1, 17):
        text = "The stairs leads {}.".format(stairType)
        print(text)
    elif diceRoll in range(17, 21):
        text = "There is an opening in the wall: a {}.".format(stairType)
        print(text)

    waitRead(text)
